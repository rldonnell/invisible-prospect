#!/usr/bin/env python3
"""
Generate professional Excel workbook from pixel_intelligence.json
Usage: python generate_excel.py <pixel_intelligence.json> <output.xlsx>
"""

import sys
import json
from datetime import datetime
from pathlib import Path
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")
LIGHT_BLUE_FILL = PatternFill(start_color="EDF5FA", end_color="EDF5FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

TIER_COLORS = {
    "HOT": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "High": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "Medium": PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid"),
    "Low": PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
}

TIER_TEXT_COLOR = Font(bold=True, color="FFFFFF")


def load_json(filepath):
    """Load JSON data from file."""
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in '{filepath}'")
        sys.exit(1)


def format_header_row(ws, row_num, columns):
    """Format header row with navy background and white text."""
    for col_num, header in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = WHITE_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_alternating_rows(ws, start_row, end_row, num_cols):
    """Apply alternating light blue shading to rows."""
    for row_num in range(start_row + 1, end_row + 1):
        if (row_num - start_row) % 2 == 0:
            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = LIGHT_BLUE_FILL


def auto_width_columns(ws, max_width=40):
    """Auto-size columns with maximum width cap."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def freeze_top_row(ws):
    """Freeze the top row."""
    ws.freeze_panes = "A2"


def create_priority_sheet(wb, data):
    """Create 'Priority List' sheet."""
    ws = wb.create_sheet("Priority List", 0)

    # Extract visitors and sort by intent score descending
    visitors = data.get("visitors", [])
    sorted_visitors = sorted(visitors, key=lambda v: v.get("intent_score", 0), reverse=True)

    columns = [
        "Intent Tier", "Score", "First Name", "Last Name", "Email", "Phone",
        "City", "State", "Age Range", "Gender", "Income", "Interests",
        "Visit Count", "First Visit", "Last Visit", "Source", "LinkedIn URL"
    ]

    format_header_row(ws, 1, columns)

    # Add data rows
    for row_num, visitor in enumerate(sorted_visitors, 2):
        tier = visitor.get("intent_tier", "Low")
        score = visitor.get("intent_score", 0)
        first_name = visitor.get("first_name", "")
        last_name = visitor.get("last_name", "")
        email = visitor.get("email", "")
        phone = visitor.get("phone", "")
        city = visitor.get("city", "")
        state = visitor.get("state", "")
        age_range = visitor.get("age_range", "")
        gender = visitor.get("gender", "")
        income = visitor.get("income_bracket", "")
        interests = ", ".join(visitor.get("interests", []))
        visit_count = visitor.get("visit_count", 0)
        first_visit = visitor.get("first_visit", "")
        last_visit = visitor.get("last_visit", "")
        source = visitor.get("source", "")
        linkedin_url = visitor.get("linkedin_url", "")

        values = [
            tier, score, first_name, last_name, email, phone, city, state,
            age_range, gender, income, interests, visit_count, first_visit,
            last_visit, source, linkedin_url
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            # Conditional formatting for Intent Tier column
            if col_num == 1:
                if tier in TIER_COLORS:
                    cell.fill = TIER_COLORS[tier]
                    cell.font = TIER_TEXT_COLOR

    apply_alternating_rows(ws, 1, len(sorted_visitors) + 1, len(columns))
    auto_width_columns(ws)
    freeze_top_row(ws)


def create_summary_sheet(wb, data):
    """Create 'Summary' sheet with key metrics."""
    ws = wb.create_sheet("Summary", 1)

    client_name = data.get("client_name", "Unknown Client")
    date_range = data.get("date_range", {})
    start_date = date_range.get("start", "N/A")
    end_date = date_range.get("end", "N/A")
    total_events = data.get("total_events", 0)
    total_visitors = data.get("total_visitors", 0)
    repeat_visitors = data.get("repeat_visitors", 0)

    # Key metrics section
    row = 1
    ws.cell(row=row, column=1).value = "Client Name:"
    ws.cell(row=row, column=2).value = client_name
    row += 1

    ws.cell(row=row, column=1).value = "Date Range:"
    ws.cell(row=row, column=2).value = f"{start_date} to {end_date}"
    row += 1

    ws.cell(row=row, column=1).value = "Total Events:"
    ws.cell(row=row, column=2).value = total_events
    row += 1

    ws.cell(row=row, column=1).value = "Total Visitors:"
    ws.cell(row=row, column=2).value = total_visitors
    row += 1

    ws.cell(row=row, column=1).value = "Repeat Visitors:"
    ws.cell(row=row, column=2).value = repeat_visitors
    row += 2

    # Tier breakdown table
    ws.cell(row=row, column=1).value = "Tier Breakdown"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    tier_headers = ["Tier", "Count", "Percentage"]
    format_header_row(ws, row, tier_headers)
    tier_row = row
    row += 1

    # Count visitors by tier
    tier_counts = defaultdict(int)
    for visitor in data.get("visitors", []):
        tier = visitor.get("intent_tier", "Low")
        tier_counts[tier] += 1

    tier_order = ["HOT", "High", "Medium", "Low"]
    for tier in tier_order:
        count = tier_counts.get(tier, 0)
        percentage = (count / total_visitors * 100) if total_visitors > 0 else 0

        ws.cell(row=row, column=1).value = tier
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{percentage:.1f}%"

        for col_num in range(1, 4):
            cell = ws.cell(row=row, column=col_num)
            cell.border = THIN_BORDER

        row += 1

    row += 1

    # Top traffic sources table
    ws.cell(row=row, column=1).value = "Top Traffic Sources"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    source_headers = ["Source", "Count"]
    format_header_row(ws, row, source_headers)
    source_row = row
    row += 1

    source_counts = Counter()
    for visitor in data.get("visitors", []):
        source = visitor.get("source", "Direct")
        source_counts[source] += 1

    for source, count in source_counts.most_common(10):
        ws.cell(row=row, column=1).value = source
        ws.cell(row=row, column=2).value = count

        for col_num in range(1, 3):
            cell = ws.cell(row=row, column=col_num)
            cell.border = THIN_BORDER

        row += 1

    row += 1

    # Top states table
    ws.cell(row=row, column=1).value = "Top States"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    state_headers = ["State", "Count"]
    format_header_row(ws, row, state_headers)
    state_row = row
    row += 1

    state_counts = Counter()
    for visitor in data.get("visitors", []):
        state = visitor.get("state", "Unknown")
        if state:
            state_counts[state] += 1

    for state, count in state_counts.most_common(10):
        ws.cell(row=row, column=1).value = state
        ws.cell(row=row, column=2).value = count

        for col_num in range(1, 3):
            cell = ws.cell(row=row, column=col_num)
            cell.border = THIN_BORDER

        row += 1

    auto_width_columns(ws)


def create_interest_sheet(wb, data):
    """Create 'Interest Breakdown' sheet."""
    ws = wb.create_sheet("Interest Breakdown", 2)

    # Get subcategory counts
    subcategory_counts = data.get("subcategory_counts", {})

    # Sort by count descending
    sorted_interests = sorted(
        subcategory_counts.items(),
        key=lambda x: x[1],
        reverse=True
    )

    columns = ["Interest/Procedure", "Visitor Count"]
    format_header_row(ws, 1, columns)

    # Add data rows
    for row_num, (interest, count) in enumerate(sorted_interests, 2):
        ws.cell(row=row_num, column=1).value = interest
        ws.cell(row=row_num, column=2).value = count

        for col_num in range(1, 3):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER

    apply_alternating_rows(ws, 1, len(sorted_interests) + 1, 2)
    auto_width_columns(ws)
    freeze_top_row(ws)


def main():
    """Main function."""
    if len(sys.argv) != 3:
        print("Usage: python generate_excel.py <pixel_intelligence.json> <output.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    # Load data
    data = load_json(input_file)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # Create sheets
    create_priority_sheet(wb, data)
    create_summary_sheet(wb, data)
    create_interest_sheet(wb, data)

    # Save workbook
    try:
        wb.save(output_file)
        print(f"Success! Excel workbook created: {output_file}")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
